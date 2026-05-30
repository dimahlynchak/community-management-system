import calendar
import io
from datetime import date
from decimal import Decimal, ROUND_DOWN

from sqlalchemy import func
from sqlalchemy.orm import Session

from app.models.charge import ChargeType, Charge
from app.models.payment import Payment
from app.models.payment_allocation import PaymentAllocation
from app.models.budget import BudgetItem
from app.models.unit import Unit
from app.schemas.finance import (
    ChargeTypeCreate, ChargeCreate, PaymentCreate, BudgetItemCreate,
)


# --- ChargeTypes ---

def create_charge_type(db: Session, community_id: int, data: ChargeTypeCreate) -> ChargeType:
    ct = ChargeType(community_id=community_id, **data.model_dump())
    db.add(ct)
    db.commit()
    db.refresh(ct)
    return ct


def get_charge_types(db: Session, community_id: int) -> list[ChargeType]:
    return db.query(ChargeType).filter(
        ChargeType.community_id == community_id,
        ChargeType.is_active == True,
    ).all()


# --- Charges ---

def calculate_amount(
    unit: Unit, charge_type: ChargeType, unit_count: int, unit_index: int = 0,
) -> Decimal:
    method = charge_type.calculation_method
    if method == "per_sqm":
        return (unit.area * charge_type.rate).quantize(Decimal("0.01"))
    if method == "fixed":
        return charge_type.rate.quantize(Decimal("0.01"))
    if method == "share":
        if unit_count == 0:
            return Decimal("0.00")
        # Largest-remainder: усі юніти отримують базову порцію, округлену вниз до копійки;
        # копієчний залишок r розподіляється по +0.01 серед перших r юнітів. Так
        # максимальна асиметрія між юнітами не перевищує 1 копійку, а сума всіх
        # нарахувань точно дорівнює rate.
        base = (charge_type.rate / unit_count).quantize(Decimal("0.01"), rounding=ROUND_DOWN)
        remainder_kopecks = int((charge_type.rate - base * unit_count) * 100)
        if unit_index < remainder_kopecks:
            return (base + Decimal("0.01")).quantize(Decimal("0.01"))
        return base
    raise ValueError(f"Unknown calculation_method: {method}")


def create_charges_for_community(
    db: Session, community_id: int, charge_type_id: int, period: str, user_id: int,
    unit_ids: list[int] | None = None,
) -> list[Charge]:
    """Нараховує charges для активних приміщень спільноти.

    Якщо unit_ids None — для всіх активних приміщень. Якщо вказано — лише для
    переліку (всі мають належати спільноті і бути активними). Метод `share`
    допускається лише для bulk-режиму (unit_ids=None), бо потребує знаменника
    рівного загальній кількості приміщень спільноти.

    Виконання атомарне: якщо хоча б один з обраних юнітів уже має нарахування
    цього типу за цей період, нічого не створюється і кидається ValueError зі
    списком конфліктних unit_id.
    """
    charge_type = db.query(ChargeType).filter(ChargeType.id == charge_type_id).first()
    if charge_type is None:
        raise ValueError("Charge type not found")
    if charge_type.community_id != community_id:
        raise ValueError("Charge type does not belong to this community")
    if not charge_type.is_active:
        raise ValueError("Charge type is not active")

    if unit_ids is None:
        # Bulk-режим: усі активні приміщення спільноти.
        units = (
            db.query(Unit)
            .filter(Unit.community_id == community_id, Unit.is_active == True)
            .all()
        )
    else:
        # Targeted-режим: лише вказані приміщення.
        if charge_type.calculation_method == "share":
            raise ValueError(
                "Method 'share' requires bulk generation for all active units; "
                "do not specify unit_ids"
            )
        units = (
            db.query(Unit)
            .filter(Unit.id.in_(unit_ids), Unit.community_id == community_id)
            .all()
        )
        found_ids = {u.id for u in units}
        missing = sorted(set(unit_ids) - found_ids)
        if missing:
            raise ValueError(f"Units not found in this community: {missing}")
        inactive = sorted(u.id for u in units if not u.is_active)
        if inactive:
            raise ValueError(
                f"Deactivated units cannot receive new charges: {inactive}"
            )

    if not units:
        return []

    # Пропускаємо юніти, у яких уже є charge цього типу за цей період: голова
    # може повторно викликати ендпоінт або донарахувати решті, не вказуючи
    # вручну тих, кому вже нараховано. Існуючі charges не змінюються.
    existing_unit_ids = {
        row[0] for row in (
            db.query(Charge.unit_id)
            .filter(
                Charge.charge_type_id == charge_type_id,
                Charge.period == period,
                Charge.unit_id.in_([u.id for u in units]),
            )
            .all()
        )
    }
    if existing_unit_ids:
        units = [u for u in units if u.id not in existing_unit_ids]
    if not units:
        return []

    unit_count = len(units)
    charges: list[Charge] = []
    for idx, unit in enumerate(units):
        amount = calculate_amount(unit, charge_type, unit_count, idx)
        charge = Charge(
            unit_id=unit.id,
            charge_type_id=charge_type_id,
            period=period,
            amount=amount,
            created_by=user_id,
        )
        db.add(charge)
        charges.append(charge)
    db.flush()
    for unit in units:
        _reallocate_unit_credit(db, unit.id)
    db.commit()
    for charge in charges:
        db.refresh(charge)
    return charges


def get_charges_by_community(
    db: Session,
    community_id: int,
    period: str | None = None,
    unit_id: int | None = None,
) -> list[Charge]:
    query = (
        db.query(Charge)
        .join(Unit)
        .filter(Unit.community_id == community_id)
        .order_by(Charge.period.desc(), Charge.id.desc())
    )
    if period:
        query = query.filter(Charge.period == period)
    return query.all()


# --- Payments ---

def _allocate_payment_fifo(db: Session, payment: Payment) -> None:
    """FIFO: allocate payment's unallocated remainder to oldest outstanding charges.

    Charges цього юніта блокуються `FOR UPDATE` до кінця транзакції: два
    паралельні платежі на той самий юніт могли б обидва прочитати однаковий
    outstanding і обидва зробити allocation, видавши allocation > charge.amount.
    З FOR UPDATE другий запит чекає першого і бачить уже оновлений залишок."""
    already_paid = (
        db.query(func.sum(PaymentAllocation.amount))
        .filter(PaymentAllocation.payment_id == payment.id)
        .scalar()
    ) or Decimal("0")
    remaining = payment.amount - already_paid
    if remaining <= Decimal("0"):
        return
    charges = (
        db.query(Charge)
        .filter(Charge.unit_id == payment.unit_id)
        .order_by(Charge.period, Charge.created_at)
        .with_for_update()
        .all()
    )
    for charge in charges:
        if remaining <= Decimal("0"):
            break
        already_allocated = (
            db.query(func.sum(PaymentAllocation.amount))
            .filter(PaymentAllocation.charge_id == charge.id)
            .scalar()
        ) or Decimal("0")
        outstanding = charge.amount - already_allocated
        if outstanding <= Decimal("0"):
            continue
        to_allocate = min(remaining, outstanding)
        db.add(PaymentAllocation(
            payment_id=payment.id,
            charge_id=charge.id,
            amount=to_allocate,
        ))
        db.flush()
        remaining -= to_allocate


def _reallocate_unit_credit(db: Session, unit_id: int) -> None:
    """Re-run FIFO for any payment of this unit that still has unallocated remainder."""
    payments = (
        db.query(Payment)
        .filter(Payment.unit_id == unit_id)
        .order_by(Payment.payment_date, Payment.created_at)
        .all()
    )
    for payment in payments:
        _allocate_payment_fifo(db, payment)


def create_payment(db: Session, data: PaymentCreate, user_id: int) -> Payment:
    payment = Payment(**data.model_dump(), created_by=user_id)
    db.add(payment)
    db.flush()  # get payment.id before allocating
    _allocate_payment_fifo(db, payment)
    db.commit()
    db.refresh(payment)
    return payment


def _recalculate_allocations_for_unit(db: Session, unit_id: int) -> None:
    """Перерозподіляє ВСІ платежі юніта FIFO заново: видаляє всі поточні
    payment_allocations цих платежів, потім послідовно по даті їхньої появи
    кладе на найстаріші непогашені нарахування цього юніта.

    Викликається після UPDATE/DELETE charge або UPDATE amount/date payment —
    коли структура боргу/оплат змінилася і попередній розподіл вже
    нерелевантний. Не комітить — це робить викликаючий код."""
    payments = (
        db.query(Payment)
        .filter(Payment.unit_id == unit_id)
        .order_by(Payment.payment_date, Payment.created_at, Payment.id)
        .all()
    )
    if not payments:
        return
    payment_ids = [p.id for p in payments]
    db.query(PaymentAllocation).filter(
        PaymentAllocation.payment_id.in_(payment_ids)
    ).delete(synchronize_session=False)
    db.flush()
    for payment in payments:
        _allocate_payment_fifo(db, payment)


def get_charge(db: Session, charge_id: int) -> Charge | None:
    return db.query(Charge).filter(Charge.id == charge_id).first()


def update_charge_amount(db: Session, charge: Charge, new_amount: Decimal) -> Charge:
    """Оновлює суму нарахування і перераховує FIFO-розподіл усіх платежів
    юніта (нова сума → інша картина непогашеного боргу). Період і тип не
    змінюються — якщо помилка в них, видаліть і створіть наново."""
    if new_amount <= Decimal("0"):
        raise ValueError("amount must be positive")
    charge.amount = new_amount.quantize(Decimal("0.01"))
    db.flush()
    _recalculate_allocations_for_unit(db, charge.unit_id)
    db.commit()
    db.refresh(charge)
    return charge


def delete_charge(db: Session, charge: Charge) -> int:
    """Видаляє нарахування. Звільнені платежі цього юніта автоматично
    перерозподіляються FIFO на інші непогашені charges. Повертає unit_id
    для аудиту після видалення."""
    unit_id = charge.unit_id
    db.query(PaymentAllocation).filter(
        PaymentAllocation.charge_id == charge.id
    ).delete(synchronize_session=False)
    db.delete(charge)
    db.flush()
    _recalculate_allocations_for_unit(db, unit_id)
    db.commit()
    return unit_id


def get_payment(db: Session, payment_id: int) -> Payment | None:
    return db.query(Payment).filter(Payment.id == payment_id).first()


def update_payment(
    db: Session,
    payment: Payment,
    new_amount: Decimal | None,
    new_date: date | None,
    new_description: str | None,
    description_explicit: bool,
) -> Payment:
    """Оновлює платіж. При зміні суми або дати — перераховує FIFO усіх платежів
    юніта (бо порядок або обсяг розподілу зміниться)."""
    structural_change = False
    if new_amount is not None:
        if new_amount <= Decimal("0"):
            raise ValueError("amount must be positive")
        payment.amount = new_amount.quantize(Decimal("0.01"))
        structural_change = True
    if new_date is not None:
        payment.payment_date = new_date
        structural_change = True
    if description_explicit:
        payment.description = new_description
    db.flush()
    if structural_change:
        _recalculate_allocations_for_unit(db, payment.unit_id)
    db.commit()
    db.refresh(payment)
    return payment


def delete_payment(db: Session, payment: Payment) -> int:
    """Видаляє платіж разом з його розподілом. Інші платежі юніта
    перерозподіляються FIFO заново. Повертає unit_id."""
    unit_id = payment.unit_id
    db.query(PaymentAllocation).filter(
        PaymentAllocation.payment_id == payment.id
    ).delete(synchronize_session=False)
    db.delete(payment)
    db.flush()
    _recalculate_allocations_for_unit(db, unit_id)
    db.commit()
    return unit_id


def get_payments_by_unit(db: Session, unit_id: int) -> list[Payment]:
    # Найновіші платежі — зверху. Без сортування Postgres повертає рядки в
    # довільному порядку, що ламає UX (пагінацію, експорти).
    return (
        db.query(Payment)
        .filter(Payment.unit_id == unit_id)
        .order_by(Payment.payment_date.desc(), Payment.id.desc())
        .all()
    )


def get_allocations_by_payment(db: Session, payment_id: int) -> list[PaymentAllocation]:
    return db.query(PaymentAllocation).filter(PaymentAllocation.payment_id == payment_id).all()


# --- Balance ---

def get_balance_for_community(db: Session, community_id: int) -> list[dict]:
    """Підсумкові баланси по всіх юнітах спільноти (включно з деактивованими —
    для збереження історичного боргу). Реалізовано двома агрегаційними
    запитами замість N+1 (по одному на юніт), що критично для спільнот з
    сотнями приміщень і тисячами charges/payments."""
    units = db.query(Unit).filter(Unit.community_id == community_id).all()
    if not units:
        return []
    unit_ids = [u.id for u in units]

    charged_by_unit = dict(
        db.query(Charge.unit_id, func.sum(Charge.amount))
        .filter(Charge.unit_id.in_(unit_ids))
        .group_by(Charge.unit_id)
        .all()
    )
    paid_by_unit = dict(
        db.query(Payment.unit_id, func.sum(Payment.amount))
        .filter(Payment.unit_id.in_(unit_ids))
        .group_by(Payment.unit_id)
        .all()
    )

    results = []
    for unit in units:
        total_charged = (charged_by_unit.get(unit.id) or Decimal("0")).quantize(Decimal("0.01"))
        total_paid = (paid_by_unit.get(unit.id) or Decimal("0")).quantize(Decimal("0.01"))
        results.append({
            "unit_id": unit.id,
            "unit_number": unit.number,
            "unit_type": unit.type,
            "total_charged": total_charged,
            "total_paid": total_paid,
            "balance": (total_paid - total_charged).quantize(Decimal("0.01")),
        })
    return results


def get_balance_for_unit(db: Session, unit: Unit) -> dict:
    total_charged = (
        db.query(func.sum(Charge.amount))
        .filter(Charge.unit_id == unit.id)
        .scalar()
    ) or Decimal("0")
    total_paid = (
        db.query(func.sum(Payment.amount))
        .filter(Payment.unit_id == unit.id)
        .scalar()
    ) or Decimal("0")
    return {
        "unit_id": unit.id,
        "unit_number": unit.number,
        "unit_type": unit.type,
        "total_charged": total_charged.quantize(Decimal("0.01")),
        "total_paid": total_paid.quantize(Decimal("0.01")),
        "balance": (total_paid - total_charged).quantize(Decimal("0.01")),
    }


# --- Penalties ---

def _last_day_of_period(period: str) -> date:
    year, month = int(period[:4]), int(period[5:7])
    return date(year, month, calendar.monthrange(year, month)[1])


def calculate_penalties(
    db: Session, community_id: int, rate: Decimal, as_of_date: date | None = None,
) -> list[dict]:
    """Розрахунок пені для всіх боржників спільноти. Оптимізовано: усі charges
    і allocations завантажуються двома запитами (без N+1 по юнітах). Для
    спільнот з сотнями приміщень і тисячами charges це різниця між сотнями
    запитів і двома."""
    if as_of_date is None:
        as_of_date = date.today()

    units = db.query(Unit).filter(Unit.community_id == community_id).all()
    if not units:
        return []
    units_by_id = {u.id: u for u in units}
    unit_ids = list(units_by_id.keys())

    charges = (
        db.query(Charge)
        .filter(Charge.unit_id.in_(unit_ids))
        .order_by(Charge.unit_id, Charge.period, Charge.created_at)
        .all()
    )
    if not charges:
        return []

    charge_ids = [c.id for c in charges]
    allocated_by_charge = dict(
        db.query(PaymentAllocation.charge_id, func.sum(PaymentAllocation.amount))
        .filter(PaymentAllocation.charge_id.in_(charge_ids))
        .group_by(PaymentAllocation.charge_id)
        .all()
    )

    results = []
    for charge in charges:
        allocated = allocated_by_charge.get(charge.id) or Decimal("0")
        debt = charge.amount - allocated
        if debt <= Decimal("0"):
            continue
        last_day = _last_day_of_period(charge.period)
        overdue_days = (as_of_date - last_day).days
        if overdue_days <= 0:
            continue
        unit = units_by_id[charge.unit_id]
        penalty = (debt * rate * overdue_days).quantize(Decimal("0.01"))
        results.append({
            "unit_id": unit.id,
            "unit_number": unit.number,
            "charge_id": charge.id,
            "period": charge.period,
            "debt": debt.quantize(Decimal("0.01")),
            "overdue_days": overdue_days,
            "rate": rate,
            "penalty": penalty,
        })
    return results


# --- My Charges ---

def get_my_charges(db: Session, unit_id: int) -> list[Charge]:
    return (
        db.query(Charge)
        .filter(Charge.unit_id == unit_id)
        .order_by(Charge.period, Charge.created_at)
        .all()
    )


# --- Export ---

def export_balance_xlsx(balance_rows: list[dict], community_name: str) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Balance"

    ws.merge_cells("A1:E1")
    ws["A1"] = f"Боргова відомість: {community_name}"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")

    headers = ["№ прим.", "Тип", "Нараховано (грн)", "Оплачено (грн)", "Баланс (грн)"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = Font(bold=True)

    for i, row in enumerate(balance_rows, 3):
        ws.cell(row=i, column=1, value=row["unit_number"])
        ws.cell(row=i, column=2, value=row["unit_type"])
        ws.cell(row=i, column=3, value=float(row["total_charged"]))
        ws.cell(row=i, column=4, value=float(row["total_paid"]))
        ws.cell(row=i, column=5, value=float(row["balance"]))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def export_balance_pdf(balance_rows: list[dict], community_name: str) -> bytes:
    from fpdf import FPDF

    pdf = FPDF()
    pdf.add_page()

    import os
    bundled_font = os.path.join(
        os.path.dirname(os.path.dirname(__file__)),
        "assets", "fonts", "DejaVuSans.ttf",
    )
    unicode_loaded = False
    for font_path in [
        bundled_font,
        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
        "/usr/share/fonts/dejavu/DejaVuSans.ttf",
        "/usr/share/fonts/TTF/DejaVuSans.ttf",
        "/Library/Fonts/DejaVuSans.ttf",
        "/Library/Fonts/Arial Unicode MS.ttf",
        "/System/Library/Fonts/Supplemental/Arial Unicode MS.ttf",
    ]:
        try:
            pdf.add_font("Unicode", fname=font_path)
            unicode_loaded = True
            break
        except Exception:
            continue

    font = "Unicode" if unicode_loaded else "Helvetica"

    def _text(s: str) -> str:
        if unicode_loaded:
            return s
        return s.encode("latin-1", errors="replace").decode("latin-1")

    pdf.set_font(font, size=14)
    pdf.cell(0, 10, _text(f"Боргова відомість: {community_name}"), align="C")
    pdf.ln(12)

    pdf.set_font(font, size=9)
    col_w = [28, 28, 42, 42, 38]
    for h, w in zip(["№ прим.", "Тип", "Нараховано", "Оплачено", "Баланс (грн)"], col_w):
        pdf.cell(w, 8, _text(h), border=1, align="C")
    pdf.ln()

    for row in balance_rows:
        vals = [
            str(row["unit_number"]),
            str(row["unit_type"]),
            str(row["total_charged"]),
            str(row["total_paid"]),
            str(row["balance"]),
        ]
        for v, w in zip(vals, col_w):
            pdf.cell(w, 7, _text(v), border=1)
        pdf.ln()

    return bytes(pdf.output())


# --- BudgetItems ---

def create_budget_item(db: Session, community_id: int, data: BudgetItemCreate, user_id: int) -> BudgetItem:
    item = BudgetItem(community_id=community_id, **data.model_dump(), created_by=user_id)
    db.add(item)
    db.commit()
    db.refresh(item)
    return item


def get_budget_items(db: Session, community_id: int, period: str | None = None) -> list[BudgetItem]:
    query = db.query(BudgetItem).filter(BudgetItem.community_id == community_id)
    if period:
        query = query.filter(BudgetItem.period == period)
    return query.all()
